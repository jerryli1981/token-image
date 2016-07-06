import wb_dict
WUBI_DICT = wb_dict.wb_dict.copy()

from openpyxl import load_workbook
wl_dict = load_workbook('Wordlist2.xlsx', use_iterators = True)
p = wl_dict.get_sheet_by_name(name = 'Sheet1')

def unicode2int(word):
    c=repr(word)
    if '\u' not in c:
        return 0
    else:
        d= c.translate(None,r"\u'")
        return int(d,16)

with open("eng2ch_dict.py", "w") as s:
	s.write("eng2ch_dict = {\n")

	for row in p.iter_rows():
		eng = row[0].value
		chi = row[1].value
		if chi == u'\uff1f' or chi == None:
			continue
		wbs = []
		for uni in chi:
			idx = unicode2int(uni)
			if idx in WUBI_DICT:
				wb = WUBI_DICT[idx]
				wbs.append(wb)

		s.write("\""+eng+"\""+":"+"\""+" ".join(wbs)+"\""+","+"\n")
		
	s.write("}")